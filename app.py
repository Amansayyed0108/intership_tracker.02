from flask import Flask, render_template, request, redirect, jsonify, send_file, session
import pandas as pd
import os
import json
import uuid

# For Excel styling
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

basedir = os.path.abspath(os.path.dirname(__file__))
templatedir = os.path.join(basedir, 'templates')
app = Flask(__name__, template_folder=templatedir)

# Secret key required to use client-side sessions securely
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "super-secret-key-change-in-production")

COLUMNS = [
    "Start Date",
    "End Date",
    "Project Task",
    "Description",
    "Tool",
    "Hours",
    "Status",
    "Supervisor"
]

# Helper function to grab unique user storage paths dynamically
def get_user_filepaths():
    if 'user_id' not in session:
        session['user_id'] = str(uuid.uuid4())
    
    # Isolate storage for this specific visitor session
    user_dir = os.path.join(basedir, "userdata", session['user_id'])
    os.makedirs(user_dir, exist_ok=True)
    
    excel_file = os.path.join(user_dir, "internshiplog.xlsx")
    supervisors_file = os.path.join(user_dir, "supervisors.json")
    return excel_file, supervisors_file

def styleexcel(excel_path):
    if not os.path.exists(excel_path):
        return

    wb = load_workbook(excel_path)
    ws = wb.active

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=11, color="000000")

    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.border = thin_border

            if col in [1, 2, 5, 6, 7, 8]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[1].height = 28
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 20

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            if len(val) > max_len:
                max_len = len(val)
        clamped_width = min(max(max_len + 4, 12), 45)
        ws.column_dimensions[col_letter].width = clamped_width

    wb.save(excel_path)

def initexcel(excel_path):
    if not os.path.exists(excel_path):
        df = pd.DataFrame(columns=COLUMNS)
        df.to_excel(excel_path, index=False)
        styleexcel(excel_path)
    else:
        try:
            df = pd.read_excel(excel_path)
            for col in COLUMNS:
                if col not in df.columns:
                    df[col] = ""
            df.to_excel(excel_path, index=False)
            styleexcel(excel_path)
        except Exception:
            df = pd.DataFrame(columns=COLUMNS)
            df.to_excel(excel_path, index=False)
            styleexcel(excel_path)

def loadsupervisors(sups_path):
    if os.path.exists(sups_path):
        try:
            with open(sups_path, 'r') as f:
                return json.load(f)
        except Exception:
            return []
    return []

def savesupervisors(sups_list, sups_path):
    with open(sups_path, 'w') as f:
        json.dump(sups_list, f)

@app.route("/")
def home():
    excel_file, supervisors_file = get_user_filepaths()
    initexcel(excel_file)
    
    df = pd.read_excel(excel_file)
    for col in ["Start Date", "End Date"]:
        if col in df.columns:
            df[col] = df[col].astype(str).replace("NaT", "").replace("nan", "")

    logs = df.to_dict(orient="records")
    logs_with_index = list(enumerate(logs))
    supervisors = loadsupervisors(supervisors_file)
    
    return render_template("index.html", logs=logs_with_index, supervisors=supervisors)

@app.route("/save", methods=["POST"])
def save():
    excel_file, _ = get_user_filepaths()
    
    row_index_str = request.form.get("row_index", "")
    start_date = request.form.get("start_date", "")
    end_date = request.form.get("end_date", "")
    task = request.form.get("task", "")
    description = request.form.get("description", "")
    tool = request.form.get("tool", "")
    hours = request.form.get("hours", "0")
    minutes = request.form.get("minutes", "00")
    status = request.form.get("status", "")
    supervisor = request.form.get("supervisor", "")

    formatted_time = f"{int(hours):02d}:{int(minutes):02d}"

    updated_row = {
        "Start Date": start_date,
        "End Date": end_date,
        "Project Task": task,
        "Description": description,
        "Tool": tool,
        "Hours": formatted_time,
        "Status": status,
        "Supervisor": supervisor
    }

    df = pd.read_excel(excel_file)

    if row_index_str != "":
        idx = int(row_index_str)
        if 0 <= idx < len(df):
            for col in COLUMNS:
                df.at[idx, col] = updated_row[col]
    else:
        df = pd.concat([df, pd.DataFrame([updated_row])], ignore_index=True)

    df.to_excel(excel_file, index=False)
    styleexcel(excel_file)
    return redirect("/")

@app.route("/delete/row/<int:index>", methods=["POST"])
def delete_row(index):
    try:
        excel_file, _ = get_user_filepaths()
        df = pd.read_excel(excel_file)
        if 0 <= index < len(df):
            df = df.drop(index).reset_index(drop=True)
            df.to_excel(excel_file, index=False)
            styleexcel(excel_file)
    except Exception as e:
        print(f"Error deleting row: {e}")
    return redirect("/")

@app.route("/add_supervisor", methods=["POST"])
def addsupervisor():
    _, supervisors_file = get_user_filepaths()
    name = request.form.get("name", "").strip()
    if name:
        sups = loadsupervisors(supervisors_file)
        if name not in sups:
            sups.append(name)
            savesupervisors(sups, supervisors_file)
    return redirect("/")

@app.route("/delete/supervisor/<string:name>", methods=["POST"])
def delete_supervisor(name):
    _, supervisors_file = get_user_filepaths()
    sups = loadsupervisors(supervisors_file)
    if name in sups:
        sups.remove(name)
        savesupervisors(sups, supervisors_file)
    return redirect("/")

@app.route("/get_last_entry")
def getlastentry():
    excel_file, _ = get_user_filepaths()
    if not os.path.exists(excel_file):
        return jsonify({})
        
    df = pd.read_excel(excel_file)
    if not df.empty:
        last_row = df.iloc[-1].fillna("").to_dict()
        return jsonify(last_row)
    return jsonify({})

@app.route("/open_excel")
def openexcel():
    try:
        excel_file, _ = get_user_filepaths()
        if os.path.exists(excel_file):
            return send_file(
                excel_file,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name="internshiplog.xlsx"
            )
        else:
            return "File not found", 404
    except Exception as e:
        return str(e), 500

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port)
